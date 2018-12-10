from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import re
import datetime
import os
from os import walk, getcwd

global_dataframe = pd.DataFrame()

mydir = getcwd()

for root, dirs, files in os.walk(mydir):
	for _file in files:
		if _file.endswith(".xlsx"):
			print(_file)
			wb = load_workbook(_file)

			sheetname = wb.sheetnames
			# print (sheetname)

			ws = wb.active
			data = ws.values

			cols = next(data)[1:]

			list_cols = list()
			list_col_nums = list()
			list_units = list()

			num_rows = ws.max_row
			print('Number of rows', num_rows)
			num_cols = ws.max_column
			print('Number of cols', num_cols)

			data_dict = dict()

			index = 1
			for item in cols:
				if item != None:
					item = item[:item.find('@')]
					list_cols.append(item)
					list_col_nums.append(index+1)

					data_dict[item] = list()

				index += 1

			units = next(data)[1:]
			index = 1

			for item in units:
				if item != None:
					if '293K' in item: item = re.sub(' 293K', '', item)
					if 'gradi C.' in item: item = re.sub('gradi C.', 'degC', item)
					if index in list_col_nums: 
						# print (item)
						list_units.append(item)
				index += 1

			# print('Columns, Index, Unit')
			# print([i for i in zip(list_cols, list_col_nums, list_units)])

			for col in list_col_nums:
				item = list_cols[list_col_nums.index(col)]
				if item != 'Time':
					# print('Putting the data in the dict', item)
					for i in range(3, num_rows):
						value = ws.cell(row=i, column=col).value
						# print(value)
						data_dict[item].append(value)

			time_list = list()
			## Get the time
			for i in range(3, num_rows):
				date = ws.cell(row=i, column=1).value
				if date != None:
					date = date.strftime('%Y-%m-%d')
					hour = ws.cell(row=i, column=2).value
					hour = datetime.datetime.strptime(hour, '%H:%M')
					hour = hour.strftime('%H:%M')
					time = datetime.datetime.strptime(date + ' ' + hour, '%Y-%m-%d %H:%M')
					# time = datetime.datetime.combine(date, hour)
					#print('Date', date)
					#print('Hour', hour)
					# print('TimeStamp', time, type(time))
					time_list.append(time)

			dataframe = pd.DataFrame(data_dict, columns = list_cols, index = time_list)
			global_dataframe = global_dataframe.combine_first(dataframe)

global_dataframe.to_csv('File_1.csv')

# # Delete the columns for the
# cols_to_delete = [i for i in range(num_cols) if i not in list_col_nums]
# print(cols_to_delete)
# 
# [ws.delete_cols(i) for i in cols_to_delete]
# 
# # Delete the rows for units since we already know them
# ws.delete_rows(1)

# header = 2
# index = 0

# num_rows = ws.max_row

# for column in ws.columns:
# 	colum
#     # for cell in row:
#         # print(cell.value)
#     # index += 1